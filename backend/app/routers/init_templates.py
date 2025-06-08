from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os

def create_po_template():
    """Crée un modèle de bon de commande Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Order"
    
    # Styles
    header_font = Font(name='Arial', size=12, bold=True)
    normal_font = Font(name='Arial', size=10)
    title_font = Font(name='Arial', size=16, bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # En-tête du document avec logo Oddnet
    ws.merge_cells('A1:F1')
    ws['A1'] = "ODDNET PURCHASE ORDER"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Informations du fournisseur
    ws['A3'] = "Supplier Information:"
    ws['A3'].font = header_font
    
    # Emplacements pour les informations du fournisseur
    ws['A4'] = "Company:"
    ws['B4'] = ""  
    
    ws['A5'] = "Address:"
    ws['B5'] = "" 
    
    ws['A6'] = "Country:"
    ws['B6'] = "" 
    
    ws['A7'] = "Contact:"
    ws['B7'] = ""  
    
    ws['A8'] = "Email:"
    ws['B8'] = ""  
    
    ws['A9'] = "Phone:"
    ws['B9'] = ""  
    
    # Informations du bon de commande
    ws['E3'] = "Purchase Order Details:"
    ws['E3'].font = header_font
    
    ws['E4'] = "PO #:"
    ws['F4'] = "" 
    
    ws['E5'] = "Date:"
    ws['F5'] = "" 
    
    ws['E6'] = "Currency:"
    ws['F6'] = "" 
    
    ws['E7'] = "ETA:"
    ws['F7'] = ""  
    
    # En-tête du tableau des articles
    ws['A11'] = "Part Number"
    ws['B11'] = "Description"
    ws['C11'] = "Quantity"
    ws['D11'] = "Unit Price"
    ws['E11'] = "Total"
    
    for cell in ['A11', 'B11', 'C11', 'D11', 'E11']:
        ws[cell].font = header_font
        ws[cell].fill = header_fill
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        ws[cell].border = thin_border
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    ws['A30'] = "Terms and Conditions:"
    ws['A30'].font = header_font
    
    ws['A31'] = "1. Please send two copies of your invoice."
    ws['A32'] = "2. Enter this order in accordance with the prices, terms, delivery method, and specifications listed above."
    ws['A33'] = "3. Please notify us immediately if you are unable to ship as specified."
    ws['A34'] = "4. Send all correspondence to: contact@oddnet.ma"
    
    # Signature
    ws['A36'] = "Authorized by:"
    ws['A36'].font = header_font
    
    ws['A38'] = "____________________"
    ws['A39'] = "ODDNET"
    
    # Sauvegarder le modèle
    template_dir = "../templates"
    os.makedirs(template_dir, exist_ok=True)
    wb.save(f"{template_dir}/po_template.xlsx")
    
    print(f"Template created at {template_dir}/po_template.xlsx")

if __name__ == "__main__":
    create_po_template()
  